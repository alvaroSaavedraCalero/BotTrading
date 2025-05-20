# reports/exporter.py

import logging
import pandas as pd
from typing import List, Tuple, Dict, Any, Union, Optional
from pandas import Timestamp

from config.config import INITIAL_CAPITAL, RISK_PERCENT, RR_RATIO

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

# TradeTuple ahora incluye PnL Neto
# TradeTuple = Tuple[str, float, float, Timestamp, Timestamp, float] # Definición local eliminada
from backtestingService.engine import TradeTuple # TradeTuple importado

# Funcion para generar un DataFrame con los resultados de los trades realizados
def generar_dataframe_resultados(historial_trades: List[TradeTuple]) -> pd.DataFrame:
    """
    Convierte el historial de trades (lista de tuplas) en un DataFrame de pandas.
    Usa el P&L neto del trade y calcula el balance acumulado real.
    """
    trades_list: List[Dict[str, Any]] = []
    current_balance: float = float(INITIAL_CAPITAL) # Nuevo balance acumulado real

    for idx, trade_info in enumerate(historial_trades, start=1):
        resultado: str; entrada: float; salida: float; fecha_entrada: Timestamp; fecha_salida: Timestamp; trade_pnl_net: float
        # Desempaquetar el nuevo TradeTuple
        resultado, entrada, salida, fecha_entrada, fecha_salida, trade_pnl_net = trade_info

        try:
            tipo: str; res: str
            tipo, res = resultado.split('_')
        except ValueError:
             logging.warning(f"Formato de resultado inesperado en trade {idx}: '{resultado}'.")
             tipo = resultado
             res = "DESCONOCIDO" # O el tipo original si no hay '_'

        current_balance += trade_pnl_net # Actualizar balance con PnL neto del trade

        # Reconstruir SL/TP para el reporte (lógica existente se mantiene)
        sl_reporte: float; tp_reporte: float
        if tipo == "LONG":
            sl_reporte = entrada * (1 - RISK_PERCENT)
            tp_reporte = entrada * (1 + RISK_PERCENT * RR_RATIO)
        elif tipo == "SHORT":
             sl_reporte = entrada * (1 + RISK_PERCENT)
             tp_reporte = entrada * (1 - RISK_PERCENT * RR_RATIO)
        else:
             sl_reporte = salida if res == 'PERDIDA' else 0.0
             tp_reporte = salida if res == 'GANANCIA' else 0.0

        stop_loss_final: float = salida if res == 'PERDIDA' else sl_reporte
        take_profit_final: float = salida if res == 'GANANCIA' else tp_reporte

        operacion: Dict[str, Any] = {
            "Operación": idx,
            "Tipo": tipo,
            "Fecha Entrada": fecha_entrada.strftime("%Y-%m-%d %H:%M") if isinstance(fecha_entrada, Timestamp) else str(fecha_entrada),
            "Fecha Salida": fecha_salida.strftime("%Y-%m-%d %H:%M") if isinstance(fecha_salida, Timestamp) else str(fecha_salida),
            "Precio Entrada": round(entrada, 5),
            "Stop Loss": round(stop_loss_final, 5),
            "Take Profit": round(take_profit_final, 5),
            "Resultado": res,
            "Profit/Loss Trade": round(trade_pnl_net, 2), # Nueva columna PnL
            "Balance Tras Operación": round(current_balance, 2) # Balance real acumulado
        }
        trades_list.append(operacion)

    # Definir columnas CON "Profit/Loss Trade" y "Balance Tras Operación" actualizado
    columnas_df: List[str] = [
        "Operación", "Tipo", "Fecha Entrada", "Fecha Salida",
        "Precio Entrada", "Stop Loss", "Take Profit", "Resultado",
        "Profit/Loss Trade", "Balance Tras Operación" # Nuevas y actualizadas columnas
    ]
    if not trades_list:
         return pd.DataFrame(columns=columnas_df)
    else:
         df = pd.DataFrame(trades_list)
         # Asegurar el orden correcto
         return df[columnas_df]


# Funcion para exportar los resultados a un archivo Excel con formato
def exportar_resultados_excel(
    trades_df: pd.DataFrame,
    balance_final_real: float, # Renombrado para claridad (este es el del backtest con comisiones)
    archivo_excel: str = "resultados_backtesting.xlsx"
) -> None:
    """
    Exporta el DataFrame de trades a Excel con formato.
    Usa el balance_final_real (del backtest) para las estadísticas finales.

    Args:
        trades_df: DataFrame generado por generar_dataframe_resultados.
        balance_final_real: El balance final del backtest (incluyendo comisiones).
        archivo_excel: Nombre del archivo Excel de salida.
    """
    if trades_df.empty:
        logging.warning("El DataFrame de trades está vacío. No se generará el archivo Excel.")
        return

    try:
        # Guardar DataFrame inicial (ahora con 9 columnas)
        trades_df.to_excel(archivo_excel, sheet_name='Backtesting', index=False, startrow=1)

        wb: Workbook = load_workbook(archivo_excel)
        ws: Worksheet = wb['Backtesting']

        # --- Formato Básico y Título ---
        ws['A1'] = 'Resultados del Backtesting'
        ws['A1'].font = Font(size=16, bold=True)
        # Ajustado a 10 columnas (A-J)
        ws.merge_cells('A1:J1') # <-- Ajustado a J1
        ws['A1'].alignment = Alignment(horizontal='center')

        # --- Formato a Cabeceras (Fila 2) ---
        header_fill: PatternFill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font: Font = Font(bold=True, color="FFFFFF")
        num_columnas: int = ws.max_column # Ahora 10 columnas
        for col_idx in range(1, num_columnas + 1):
            cell = ws.cell(row=2, column=col_idx)
            if cell:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

        # --- Formato Condicional a Filas de Datos (desde Fila 3) ---
        if ws.max_row >= 3:
             for row_idx in range(3, ws.max_row + 1):
                # Ajustar índices de columna para formato
                tipo_cell = ws.cell(row=row_idx, column=2)
                resultado_cell = ws.cell(row=row_idx, column=8) # Col H
                profit_loss_cell = ws.cell(row=row_idx, column=9) # Nueva Col I (P/L)
                balance_cell = ws.cell(row=row_idx, column=10)    # Nueva Col J (Balance)
                precio_entrada_cell = ws.cell(row=row_idx, column=5)
                stop_loss_cell = ws.cell(row=row_idx, column=6)
                take_profit_cell = ws.cell(row=row_idx, column=7)

                # Formato Tipo (sin cambios en lógica, solo celda)
                tipo_value: Any = tipo_cell.value
                if tipo_value:
                    tipo_str: str = str(tipo_value).upper()
                    if tipo_str == "LONG":
                        tipo_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        tipo_cell.font = Font(color="006100", bold=True)
                    elif tipo_str == "SHORT":
                        tipo_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        tipo_cell.font = Font(color="9C0006", bold=True)
                    tipo_cell.alignment = Alignment(horizontal='center')

                # Formato Celdas Numéricas (Precios)
                for celda_num in [precio_entrada_cell, stop_loss_cell, take_profit_cell]:
                    if celda_num:
                        celda_num.number_format = '#,##0.00000'
                        celda_num.alignment = Alignment(horizontal='center')

                # Formato Fuente Precios
                if precio_entrada_cell: precio_entrada_cell.font = Font(color="1F497D", bold=True)
                if stop_loss_cell: stop_loss_cell.font = Font(color="9C0006", bold=True)
                if take_profit_cell: take_profit_cell.font = Font(color="006100", bold=True)

                # Formato Resultado
                resultado_value: Any = resultado_cell.value
                if resultado_value == 'GANANCIA':
                    resultado_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    resultado_cell.font = Font(color="006100", bold=True)
                elif resultado_value == 'PERDIDA':
                    resultado_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    resultado_cell.font = Font(color="9C0006", bold=True)
                if resultado_cell: resultado_cell.alignment = Alignment(horizontal='center')

                # Formato para Profit/Loss Trade (Columna I)
                if profit_loss_cell:
                    profit_loss_cell.number_format = '#,##0.00 €'
                    profit_loss_cell.alignment = Alignment(horizontal='right')
                    # Colorear P/L
                    pnl_value: Any = profit_loss_cell.value
                    if isinstance(pnl_value, (int, float)):
                        if pnl_value > 0:
                            profit_loss_cell.font = Font(color="006100", bold=True) # Verde para positivo
                        elif pnl_value < 0:
                            profit_loss_cell.font = Font(color="9C0006", bold=True) # Rojo para negativo

                # Formato para balance_cell (Columna J)
                if balance_cell:
                    balance_cell.number_format = '#,##0.00 €'
                    balance_cell.alignment = Alignment(horizontal='right')

        # --- Ajuste automático del ancho de columnas ---
        # (La lógica existente debería funcionar para 10 columnas)
        for col_cells in ws.columns:
            max_length: int = 0
            col_index: Optional[int] = col_cells[0].column
            if col_index is not None:
                column_letter: str = get_column_letter(col_index)
                for cell in col_cells:
                    try:
                        if cell.value is not None:
                            cell_len: int = len(str(cell.value))
                            if cell_len > max_length:
                                max_length = cell_len
                    except Exception as e_cell_len:
                         logging.debug(f"No se pudo obtener longitud para celda desconocida: {e_cell_len}")
                         pass
                adjusted_width: float = max(10.0, (max_length + 2) * 1.2)
                ws.column_dimensions[column_letter].width = adjusted_width
            else:
                logging.warning("Se encontró una columna sin índice numérico válido al ajustar anchos.")

        # --- Añadir Estadísticas ---
        ultima_fila: int = ws.max_row + 2
        tp_count: int = int((trades_df['Resultado'] == 'GANANCIA').sum())
        sl_count: int = int((trades_df['Resultado'] == 'PERDIDA').sum())
        total_trades: int = len(trades_df)

        # Usar balance_final_real (del backtest con comisiones) para estadísticas
        rentabilidad_real: float = 0.0
        if INITIAL_CAPITAL != 0 and total_trades > 0: # Evitar división por cero
             rentabilidad_real = ((balance_final_real - INITIAL_CAPITAL) / INITIAL_CAPITAL) * 100

        estadisticas: List[List[Any]] = [
            ["📈 Estadísticas del rendimiento", ""],
            ["Trades realizados:", total_trades],
            ["Take Profit alcanzado:", f"{tp_count} trades ({(tp_count / total_trades * 100):.2f}%)" if total_trades > 0 else "0 trades (N/A)"],
            ["Stop Loss alcanzado:", f"{sl_count} trades ({(sl_count / total_trades * 100):.2f}%)" if total_trades > 0 else "0 trades (N/A)"],
            ["Balance Inicial:", f"{INITIAL_CAPITAL:.2f} €"],
            ["Balance Final:", f"{balance_final_real:.2f} €"], # Usar balance_final_real
            ["Rentabilidad total:", f"{rentabilidad_real:.2f}%"] # Usar rentabilidad_real
        ]
        # (Bucle para escribir estadísticas idéntico al anterior)
        for idx, stats_row in enumerate(estadisticas, start=ultima_fila):
            etiqueta: str = str(stats_row[0])
            valor: Any = stats_row[1]
            ws[f'A{idx}'] = etiqueta
            ws[f'B{idx}'] = valor
            ws[f'A{idx}'].font = Font(bold=True, size=11)
            if isinstance(valor, (int, float)):
                 ws[f'B{idx}'].number_format = '#,##0.00'
                 ws[f'B{idx}'].alignment = Alignment(horizontal='right')
            else:
                 ws[f'B{idx}'].alignment = Alignment(horizontal='left')

        wb.save(archivo_excel)
        logging.info(f"✅ Excel generado exitosamente en: '{archivo_excel}'")

    except Exception as e:
        logging.error(f"❌ Error al exportar resultados a Excel: {e}", exc_info=True)


# Función para mostrar los resultados de forma tabular y clara en logs
def mostrar_resultados(
    balance_final_real: float, # Renombrado para claridad (este es el del backtest con comisiones)
    historial_trades: List[TradeTuple]
) -> None:
    """
    Muestra un resumen de los resultados del backtesting en los logs.
    Usa el balance_final_real (del backtest) para las estadísticas finales.

    Args:
        balance_final_real: El balance final del backtest (incluyendo comisiones).
        historial_trades: Lista de tuplas representando cada trade.
    """
    logging.info("\n--- Resultados del Backtesting ---")
    # Generar DF CON PnL y balance real
    trades_df: pd.DataFrame = generar_dataframe_resultados(historial_trades)

    if trades_df.empty:
        logging.info("No se ejecutaron trades durante el backtesting.")
    else:
        logging.info("Detalle de Operaciones:")
        try:
             # Ajustar ancho para 10 columnas
             with pd.option_context('display.max_rows', None, 'display.max_columns', None, 'display.width', 1200):
                logging.info(f"\n{trades_df.to_string(index=False)}\n")
        except Exception as e_print:
             logging.error(f"Error al formatear DataFrame para log: {e_print}")
             logging.info(trades_df)

    # Calcular estadísticas
    total_trades: int = len(historial_trades)
    tp_count: int = 0
    sl_count: int = 0
    if not trades_df.empty:
         tp_count = int((trades_df['Resultado'] == 'GANANCIA').sum())
         sl_count = int((trades_df['Resultado'] == 'PERDIDA').sum())

    porcentaje_tp: float = (tp_count / total_trades * 100) if total_trades > 0 else 0.0
    porcentaje_sl: float = (sl_count / total_trades * 100) if total_trades > 0 else 0.0

    # Usar balance_final_real para calcular rentabilidad
    rentabilidad_real: float = 0.0
    if INITIAL_CAPITAL != 0: # Evitar división por cero
        rentabilidad_real = ((balance_final_real - INITIAL_CAPITAL) / INITIAL_CAPITAL) * 100

    logging.info("\n📈 Estadísticas del rendimiento:")
    logging.info(f"- Trades realizados: {total_trades}")
    logging.info(f"- Take Profit alcanzado: {tp_count} trades ({porcentaje_tp:.2f}%)")
    logging.info(f"- Stop Loss alcanzado: {sl_count} trades ({porcentaje_sl:.2f}%)")
    logging.info(f"- Balance inicial: {INITIAL_CAPITAL:.2f} €")
    logging.info(f"- Balance final: {balance_final_real:.2f} €") # Usar balance_final_real
    logging.info(f"- Rentabilidad total: {rentabilidad_real:.2f}%") # Usar rentabilidad_real
    logging.info("------------------------------------")
    logging.warning("DISCLAIMER: Backtest P&L is based on a direct percentage-of-balance risk model. Live trading P&L will depend on explicit position sizing (via calcular_cantidad), actual market fill prices, and slippage, and may differ significantly from backtest results.")